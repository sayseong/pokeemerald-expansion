import re
import os
from openpyxl import load_workbook

# 文件路径
c_file_path = os.path.dirname(os.path.abspath(__file__))+"/../src/battle_message.c"
xlsx_file_path = os.path.dirname(os.path.abspath(__file__))+"/src/战斗文本.xlsx"

# 读取 xlsx 文件并解析为字典
def load_translations(xlsx_file_path):
    translations = {}
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):  # 从第二行开始读取
        if row[0] and row[2]:  # 确保第一列和第三列都有内容
            key = row[0].strip().lower()  # 忽略大小写
            value = row[2].strip()
            translations[key] = value
    return translations

# 替换 C 文件中的内容
def replace_c_file(c_file_path, translations):
    with open(c_file_path, "r", encoding="utf-8") as c_file:
        lines = c_file.readlines()

    updated_lines = []
    for line in lines:
        # 跳过注释掉的行
        if line.strip().startswith("//"):
            updated_lines.append(line)
            continue

        # 查找双引号之间的内容
        match = re.search(r'"(.*?)"', line)
        if match:
            original_text = match.group(1)
            # 检查是否有匹配的 key（忽略大小写）
            for key, replacement in translations.items():
                # 确保 key 匹配且后续字符是 "[" 或 "]"
                key_index = line.lower().find(key)
                if key_index != -1 and (key_index + len(key) == len(line) or line[key_index + len(key)] in "[]"):
                    # 替换双引号中的内容
                    escaped_replacement = replacement.replace('"', r'\"')  # 保留转义符
                    line = line[:match.start(1)] + escaped_replacement + line[match.end(1):]
                    break
        updated_lines.append(line)

    # 将更新后的内容写回 C 文件
    with open(c_file_path, "w", encoding="utf-8") as c_file:
        c_file.writelines(updated_lines)

# 主函数
def main():
    translations = load_translations(xlsx_file_path)
    replace_c_file(c_file_path, translations)
    print("替换完成！")

if __name__ == "__main__":
    main()