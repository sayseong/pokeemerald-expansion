import re
import openpyxl
import os

def load_replacement_dict(xlsx_file):
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb['Sheet1']
    replacement_dict = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        key = row[0].value
        value = row[4].value
        if key and value:
            replacement_dict[key] = value
    return replacement_dict

def load_replacement_dict2(xlsx_file):
    wb = openpyxl.load_workbook(xlsx_file)
    sheet = wb['Sheet2']
    replacement_dict = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=5):
        key = row[0].value
        value = row[1].value
        if key and value:
            replacement_dict[key] = value
    return replacement_dict

# 替换 C 文件中的双引号内容
def replace_in_c_file(c_file, replacement_dict):
    with open(c_file, 'r', encoding='utf-8') as file:
        content = file.read()

    # 匹配双引号之间的内容
    def replace_match(match):
        original_text = match.group(1)
        # 如果匹配的内容在字典中，替换为字典的值
        return f'"{replacement_dict.get(original_text, original_text)}"'

    # 使用正则表达式替换双引号中的内容
    updated_content = re.sub(r'"(.*?)"', replace_match, content)

    # 将修改后的内容写回文件
    with open(c_file, 'w', encoding='utf-8') as file:
        file.write(updated_content)

# 主函数
def main():
    current_folder = os.path.dirname(os.path.abspath(__file__))
    xlsx_file = current_folder+"/src/debug文本.xlsx"  # Excel 文件路径
    c_file = current_folder+"/../src/battle_debug.c"       # C 文件路径

    # 加载替换字典
    replacement_dict = load_replacement_dict(xlsx_file)

    # 替换 C 文件中的内容
    replace_in_c_file(c_file, replacement_dict)

    # c_file2 = current_folder+"/../src/debug.c"       # C 文件路径

    # replacement_dict = load_replacement_dict2(xlsx_file)

    # # 替换 C 文件中的内容
    # replace_in_c_file(c_file2, replacement_dict)

if __name__ == '__main__':
    main()