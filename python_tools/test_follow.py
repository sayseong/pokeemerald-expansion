import openpyxl
import re

# 文件路径
excel_path = r"c:\Users\Nox\Documents\GitHub\pokeemerald-expansion-Chinese\python_tools\src\精灵跟随.xlsx"
c_file_path = r"c:\Users\Nox\Documents\GitHub\pokeemerald-expansion-Chinese\src\follower_helper.c"

# 读取 Excel 文件
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

# 提取 Excel 数据(第一列和第三列)
replacement_map = {}
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
    variable_name = row[0].value  # 第一列
    replacement_text = row[2].value  # 第三列
    if variable_name and replacement_text:
        # 将 \n 转义为 \\n，确保在代码中保持原样
        replacement_map[variable_name] = replacement_text.replace("\\", "\\\\").replace("\n", "\\n")

# 读取 C 文件内容
with open(c_file_path, "r", encoding="utf-8") as file:
    c_code = file.read()

# 替换变量对应的字符串
for variable, new_text in replacement_map.items():
    # 匹配变量对应的字符串定义
    pattern = rf'static const u8 {variable}\[\] = _\(".*?"\);'
    replacement = rf'static const u8 {variable}[] = _("{new_text}");'
    c_code, num_replacements = re.subn(pattern, replacement, c_code, flags=re.DOTALL)

    # 输出调试信息
    if num_replacements == 0:
        print(f"未匹配到变量: {variable}")
    else:
        print(f"替换成功: {variable} -> {new_text}")

# 保存修改后的 C 文件
with open(c_file_path, "w", encoding="utf-8") as file:
    file.write(c_code)

print("替换完成！")